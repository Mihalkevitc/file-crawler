import os
import csv
import zipfile
import rarfile
import py7zr
import tempfile
from pathlib import Path
import chardet
from docx import Document
from openpyxl import load_workbook
import PyPDF2

class FileCrawler:
    """
    Краулер для рекурсивного обхода папок и извлечения текста из файлов.
    Поддерживает: doc, docx, xls, xlsx, pdf, txt, zip, rar, 7z
    """
    
    def __init__(self, root_dir='data', output_file='output/file_index.csv'):
        self.root_dir = root_dir
        self.output_file = output_file
        self.results = []
        
    def scan(self, path):
        """Рекурсивно обходим директорию"""
        for item in Path(path).iterdir():
            if item.is_dir():
                self.scan(item)
            else:
                self.process(item)
    
    def process(self, file_path):
        """Обрабатываем один файл: определяем тип и извлекаем текст"""
        ext = file_path.suffix.lower()
        info = {
            'file_path': str(file_path),
            'file_name': file_path.name,
            'file_type': ext,
            'file_size': file_path.stat().st_size,
            'content': ''
        }
        
        # Текстовые файлы
        if ext == '.txt':
            with open(file_path, 'rb') as f:
                raw = f.read()
                enc = chardet.detect(raw)['encoding'] or 'utf-8'
                info['content'] = raw.decode(enc, errors='ignore')
        
        # Word документы
        elif ext in ['.doc', '.docx']:
            try:
                doc = Document(file_path)
                info['content'] = '\n'.join([p.text for p in doc.paragraphs])
            except:
                # Если не открывается как docx, пробуем как простой текст
                with open(file_path, 'rb') as f:
                    raw = f.read()
                    enc = chardet.detect(raw)['encoding'] or 'utf-8'
                    info['content'] = raw.decode(enc, errors='ignore')
        
        # Excel таблицы
        elif ext in ['.xls', '.xlsx']:
            try:
                wb = load_workbook(file_path, data_only=True)
                text = []
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        text.append(' '.join([str(c) for c in row if c]))
                info['content'] = '\n'.join(text)
            except:
                with open(file_path, 'rb') as f:
                    raw = f.read()
                    enc = chardet.detect(raw)['encoding'] or 'utf-8'
                    info['content'] = raw.decode(enc, errors='ignore')
        
        # PDF файлы
        elif ext == '.pdf':
            try:
                with open(file_path, 'rb') as f:
                    pdf = PyPDF2.PdfReader(f)
                    text = []
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text.append(page_text)
                    info['content'] = '\n'.join(text)
            except:
                with open(file_path, 'rb') as f:
                    raw = f.read()
                    enc = chardet.detect(raw)['encoding'] or 'utf-8'
                    info['content'] = raw.decode(enc, errors='ignore')
        
        # Архивы - распаковываем и обрабатываем содержимое
        elif ext in ['.zip', '.rar', '.7z']:
            with tempfile.TemporaryDirectory() as tmp:
                try:
                    if ext == '.zip':
                        with zipfile.ZipFile(file_path, 'r') as zf:
                            zf.extractall(tmp)
                    elif ext == '.rar':
                        with rarfile.RarFile(file_path, 'r') as rf:
                            rf.extractall(tmp)
                    elif ext == '.7z':
                        with py7zr.SevenZipFile(file_path, 'r') as sz:
                            sz.extractall(tmp)
                    
                    for root, _, files in os.walk(tmp):
                        for f in files:
                            self.process(Path(root) / f)
                except:
                    pass  # если архив битый - пропускаем
            return  # сам архив не добавляем в результаты
        
        self.results.append(info)
    
    def save(self):
        """Сохраняем результаты в CSV файл"""
        os.makedirs(os.path.dirname(self.output_file), exist_ok=True)
        with open(self.output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['file_path', 'file_name', 
                                                    'file_type', 'file_size', 'content'])
            writer.writeheader()
            writer.writerows(self.results)
    
    def run(self):
        """Запускаем сканирование и сохранение"""
        self.scan(Path(self.root_dir))
        self.save()

if __name__ == "__main__":
    FileCrawler().run()